"""
讀取 NIfTI (.nii / .nii.gz) 檔案的 header 資訊，並可匯出 Excel。

用法：
    # 讀指定資料夾內所有 .nii.gz
    python IXI\\read_nii_header.py  --dir IXI\\IXI_preprocessed\\nii

    # 讀指定檔案
    python IXI\\read_nii_header.py  IXI\\atlas_mni152_09c.nii.gz

    # 資料夾 + 額外檔案，匯出 Excel
    python IXI\\read_nii_header.py  --dir IXI\\IXI_preprocessed\\nii  IXI\\atlas_mni152_09c.nii.gz  --excel output.xlsx

    # 顯示完整 header
    python IXI\\read_nii_header.py  --dir IXI\\IXI_preprocessed\\nii  --full
"""

import os
import sys
import glob
import argparse
import numpy as np

parser = argparse.ArgumentParser(description='讀取 NIfTI header 資訊')
parser.add_argument('files', nargs='*', help='.nii 或 .nii.gz 檔案路徑（可多個，支援 wildcard）')
parser.add_argument('--dir', type=str, default=None,
                    help='指定資料夾，自動讀取內部所有 .nii / .nii.gz')
parser.add_argument('--full', action='store_true', help='顯示完整 header（預設只顯示重點）')
parser.add_argument('--excel', type=str, default=None, metavar='PATH',
                    help='匯出 Excel 檔案路徑（如 header_report.xlsx）')
args = parser.parse_args()

try:
    import nibabel as nib
except ImportError:
    print("需要 nibabel：pip install nibabel")
    sys.exit(1)

# ── 收集檔案 ─────────────────────────────────────────────────────────
all_files = []

# 從 --dir 收集
if args.dir:
    dir_path = os.path.normpath(args.dir)
    if not os.path.isdir(dir_path):
        print(f"❌ 資料夾不存在：{dir_path}")
        sys.exit(1)
    for ext in ['*.nii', '*.nii.gz']:
        all_files.extend(glob.glob(os.path.join(dir_path, ext)))

# 從 positional args 收集（支援 wildcard）
for pattern in (args.files or []):
    expanded = glob.glob(pattern)
    if expanded:
        all_files.extend(expanded)
    elif os.path.exists(pattern):
        all_files.append(pattern)
    else:
        print(f"找不到：{pattern}")

# 過濾 + 去重 + 排序
all_files = [f for f in all_files if f.endswith('.nii') or f.endswith('.nii.gz')]
all_files = sorted(set(all_files))

if not all_files:
    print("沒有找到任何 .nii / .nii.gz 檔案")
    print("用法：python read_nii_header.py --dir <資料夾>  或  python read_nii_header.py <檔案...>")
    sys.exit(1)

print(f"共 {len(all_files)} 個檔案\n")
print("=" * 90)

# ── 讀取 header ──────────────────────────────────────────────────────
records = []

for i, filepath in enumerate(all_files):
    img = nib.load(filepath)
    hdr = img.header
    affine = img.affine

    name = os.path.basename(filepath)
    shape = img.shape
    spacing = tuple(round(float(s), 4) for s in hdr.get_zooms()[:3])
    orientation = nib.aff2axcodes(affine)
    dtype = str(hdr.get_data_dtype())

    # voxel size from affine
    voxel_sizes = np.sqrt(np.sum(affine[:3, :3] ** 2, axis=0))
    voxel_str = tuple(round(float(v), 4) for v in voxel_sizes)

    # origin
    origin = affine[:3, 3]
    origin_str = tuple(round(float(o), 2) for o in origin)

    # direction
    direction = affine[:3, :3] / voxel_sizes[np.newaxis, :]
    is_identity = np.allclose(direction, np.eye(3), atol=0.01)
    is_diag = np.allclose(direction, np.diag(np.diag(direction)), atol=0.01)

    if is_identity:
        dir_desc = 'Identity'
    elif is_diag:
        diag = np.diag(direction)
        dir_desc = f'Diagonal {tuple(round(float(d), 2) for d in diag)}'
    else:
        dir_desc = '有旋轉分量'

    import datetime
    import os
    
    # 印出 (MATLAB style)
    print(f"\n[{i+1}/{len(all_files)}]")
    

    
    # Get metadata
    filepath_abs = os.path.abspath(filepath)
    filemoddate = datetime.datetime.fromtimestamp(os.path.getmtime(filepath)).strftime('%d-%b-%Y %H:%M:%S')
    filesize = os.path.getsize(filepath)
    version = hdr['magic'].tobytes().decode('utf-8', 'ignore').strip(' \x00') if 'magic' in hdr else 'Unknown'
    if version == 'n+1': version = 'NIfTI1'
    desc = hdr['descrip'].tobytes().decode('utf-8', 'ignore').strip(' \x00') if 'descrip' in hdr else ''
    
    shape_str = ' '.join([str(s) for s in shape])
    sp_str = ' '.join([f"{s:.4f}" for s in spacing])
    bitpix = int(hdr['bitpix']) if 'bitpix' in hdr else 0
    
    xyzt_units = hdr.get_xyzt_units()
    space_unit = xyzt_units[0] if xyzt_units and len(xyzt_units) > 0 else 'Unknown'
    time_unit = xyzt_units[1] if xyzt_units and len(xyzt_units) > 1 else 'Unknown'
    space_unit_str = space_unit.capitalize() if space_unit else 'Unknown'
    time_unit_str = time_unit.capitalize() if time_unit else 'Unknown'
    if space_unit_str == 'Mm': space_unit_str = 'Millimeter'
    if time_unit_str == 'Sec': time_unit_str = 'Second'
    
    add_off = float(hdr['scl_inter']) if 'scl_inter' in hdr and not np.isnan(hdr['scl_inter']) else 0
    mul_scale = float(hdr['scl_slope']) if 'scl_slope' in hdr and not np.isnan(hdr['scl_slope']) else 1
    toffset = float(hdr['toffset']) if 'toffset' in hdr else 0
    
    # slice_code mapping
    slice_code = int(hdr['slice_code']) if 'slice_code' in hdr else 0
    slice_code_str = 'Unknown'
    if slice_code == 1: slice_code_str = 'Sequential (Increasing)'
    elif slice_code == 2: slice_code_str = 'Sequential (Decreasing)'
    elif slice_code == 3: slice_code_str = 'Interleaved (Increasing)'
    elif slice_code == 4: slice_code_str = 'Interleaved (Decreasing)'
    
    dim_info = int(hdr['dim_info']) if 'dim_info' in hdr else 0
    freq_dim = dim_info & 0x03
    phase_dim = (dim_info >> 2) & 0x03
    spatial_dim = (dim_info >> 4) & 0x03
    
    cal_min = float(hdr['cal_min']) if 'cal_min' in hdr else 0
    cal_max = float(hdr['cal_max']) if 'cal_max' in hdr else 0
    
    qform_code = int(hdr['qform_code']) if 'qform_code' in hdr else 0
    sform_code = int(hdr['sform_code']) if 'sform_code' in hdr else 0
    transform_name = 'Sform' if sform_code > 0 else ('Qform' if qform_code > 0 else 'Unknown')
    
    qfactor = float(hdr['pixdim'][0]) if 'pixdim' in hdr else 1
    if qfactor not in (1, -1): qfactor = 1

    print(f"{'Filename':>25}: '{filepath_abs}'")
    print(f"{'Filemoddate':>25}: '{filemoddate}'")
    print(f"{'Filesize':>25}: {filesize}")
    print(f"{'Version':>25}: '{version}'")
    print(f"{'Description':>25}: '{desc}'")
    print(f"{'ImageSize':>25}: [{shape_str}]")
    print(f"{'PixelDimensions':>25}: [{sp_str}]")
    print(f"{'Datatype':>25}: '{dtype}'")
    print(f"{'BitsPerPixel':>25}: {bitpix}")
    print(f"{'SpaceUnits':>25}: '{space_unit_str}'")
    print(f"{'TimeUnits':>25}: '{time_unit_str}'")
    print(f"{'AdditiveOffset':>25}: {add_off:g}")
    print(f"{'MultiplicativeScaling':>25}: {mul_scale:g}")
    print(f"{'TimeOffset':>25}: {toffset:g}")
    print(f"{'SliceCode':>25}: '{slice_code_str}'")
    print(f"{'FrequencyDimension':>25}: {freq_dim}")
    print(f"{'PhaseDimension':>25}: {phase_dim}")
    print(f"{'SpatialDimension':>25}: {spatial_dim}")
    print(f"{'DisplayIntensityRange':>25}: [{cal_min:g} {cal_max:g}]")
    print(f"{'TransformName':>25}: '{transform_name}'")
    print(f"{'Transform':>25}: [1x1 affine3d]")
    print(f"{'Qfactor':>25}: {qfactor:g}")
    print(f"{'raw':>25}: [1x1 struct]")

    # 將所有計算出的資訊存入 records 準備寫入 Excel
    records.append({
        '檔案名稱': name,
        'Shape': str(shape),
        'Spacing_0': spacing[0] if len(spacing) > 0 else None,
        'Spacing_1': spacing[1] if len(spacing) > 1 else None,
        'Spacing_2': spacing[2] if len(spacing) > 2 else None,
        'Orientation': str(orientation),
        'Direction': dir_desc,
        'Filename': filepath_abs,
        'Filemoddate': filemoddate,
        'Filesize': filesize,
        'Version': version,
        'Description': desc,
        'ImageSize': f"[{shape_str}]",
        'PixelDimensions': f"[{sp_str}]",
        'Datatype': dtype,
        'BitsPerPixel': bitpix,
        'SpaceUnits': space_unit_str,
        'TimeUnits': time_unit_str,
        'AdditiveOffset': add_off,
        'MultiplicativeScaling': mul_scale,
        'TimeOffset': toffset,
        'SliceCode': slice_code_str,
        'FrequencyDimension': freq_dim,
        'PhaseDimension': phase_dim,
        'SpatialDimension': spatial_dim,
        'DisplayIntensityRange': f"[{cal_min:g} {cal_max:g}]",
        'TransformName': transform_name,
        'Transform': '[1x1 affine3d]',
        'Qfactor': qfactor,
        'raw': '[1x1 struct]',
    })

    if args.full:
        print(f"\n  --- 完整 Header ---")
        print(f"  {hdr}")


    print("-" * 90)

# ── 總結表 ────────────────────────────────────────────────────────────
if len(all_files) > 1:
    print(f"\n{'='*90}")
    print(f"  總結（{len(all_files)} 個檔案）")
    print(f"{'='*90}")
    print(f"\n  {'檔案名稱':<45} {'Shape':<22} {'Spacing':<22} {'Orient'}")
    print(f"  {'-'*43} {'-'*20} {'-'*20} {'-'*10}")

    for rec in records:
        print(f"  {rec['檔案名稱']:<45} {rec['Shape']:<22} "
              f"({rec['Spacing_0']}, {rec['Spacing_1']}, {rec['Spacing_2']})  "
              f"{rec['Orientation']}")

    # 一致性檢查
    shapes = set(rec['Shape'] for rec in records)
    spacings = set((rec['Spacing_0'], rec['Spacing_1'], rec['Spacing_2']) for rec in records)
    orients = set(rec['Orientation'] for rec in records)
    directions = set(rec['Direction'] for rec in records)

    print(f"\n  一致性檢查：")
    print(f"    Shape 一致       : {'✅ ' + str(shapes.pop()) if len(shapes) == 1 else '❌ 有 ' + str(len(shapes)) + ' 種'}")
    print(f"    Spacing 一致     : {'✅ ' + str(spacings.pop()) if len(spacings) == 1 else '❌ 有 ' + str(len(spacings)) + ' 種'}")
    print(f"    Orientation 一致 : {'✅ ' + str(orients.pop()) if len(orients) == 1 else '❌ 有 ' + str(len(orients)) + ' 種'}")
    print(f"    Direction 一致   : {'✅ ' + str(directions.pop()) if len(directions) == 1 else '❌ 有 ' + str(len(directions)) + ' 種'}")

# ── 匯出 Excel ────────────────────────────────────────────────────────
if args.excel:
    try:
        import openpyxl
    except ImportError:
        print("\n⚠️  需要 openpyxl 才能匯出 Excel：pip install openpyxl")
        print("   改用 CSV 匯出...")
        csv_path = args.excel.replace('.xlsx', '.csv').replace('.xls', '.csv')
        if not csv_path.endswith('.csv'):
            csv_path += '.csv'
        import csv
        with open(csv_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.DictWriter(f, fieldnames=records[0].keys())
            writer.writeheader()
            writer.writerows(records)
        print(f"✅ 已匯出 CSV：{csv_path}")
        sys.exit(0)

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = 'NIfTI Headers'

    # 欄位定義
    columns = list(records[0].keys())

    # 標題列樣式
    header_font = Font(bold=True, color='FFFFFF', size=10)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # 寫標題
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # 寫資料
    alt_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
    for row_idx, rec in enumerate(records, 2):
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=rec[col_name])
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    # 自動調整欄寬
    for col_idx, col_name in enumerate(columns, 1):
        max_len = len(str(col_name))
        for row_idx in range(2, len(records) + 2):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 3, 40)

    # 凍結首列
    ws.freeze_panes = 'A2'

    # 儲存
    excel_path = args.excel
    if not excel_path.endswith(('.xlsx', '.xls')):
        excel_path += '.xlsx'
    wb.save(excel_path)
    print(f"\n✅ 已匯出 Excel：{excel_path}（{len(records)} 筆）")
